"""Routes for employee management, attendance, payroll, and CV handling."""
import sqlite3
from flask import (
    Blueprint, render_template, request, redirect, url_for,
    session, flash, send_from_directory, send_file, jsonify, Response
)
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from db import get_db_connection
from leave_management import (
    calculate_leave_days, get_employee_leave_balance, create_leave_balance_for_employee,
    submit_leave_request, approve_leave_request, reject_leave_request,
    get_employee_leave_requests, get_all_leave_requests, get_pending_leave_requests,
    get_employee_leave_balances
)
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
import openpyxl
import os
import csv
import io
from validators import validate_employee_data
from config import BASE_DIR

pdfmetrics.registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))

employees_bp = Blueprint('employees', __name__)

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads", "cvs")


@employees_bp.route('/employees')
def employees():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    keyword = request.args.get('keyword', '')
    try:
        per_page = int(request.args.get('per_page', 20))
        if per_page not in (10, 20, 50):
            per_page = 20
    except ValueError:
        per_page = 20
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1

    offset = (page - 1) * per_page

    conn = get_db_connection()
    if keyword:
        like = '%' + keyword + '%'
        employees = conn.execute(
            "SELECT * FROM employees WHERE (name LIKE ? OR email LIKE ? OR phone LIKE ? OR COALESCE(address,'') LIKE ? OR COALESCE(department,'') LIKE ?) LIMIT ? OFFSET ?",
            (like, like, like, like, like, per_page, offset)
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM employees WHERE (name LIKE ? OR email LIKE ? OR phone LIKE ? OR COALESCE(address,'') LIKE ? OR COALESCE(department,'') LIKE ? )",
            (like, like, like, like, like)
        ).fetchone()['c']
    else:
        employees = conn.execute(
            "SELECT * FROM employees LIMIT ? OFFSET ?",
            (per_page, offset)
        ).fetchall()
        total = conn.execute(
            "SELECT COUNT(*) AS c FROM employees"
        ).fetchone()['c']

    conn.close()

    pages = (total + per_page - 1) // per_page if total else 1

    return render_template(
        'employee_list.html',
        employees=employees,
        keyword=keyword,
        page=page,
        pages=pages,
        total=total,
        per_page=per_page
    )


@employees_bp.route('/employee/<int:emp_id>')
def employee_details(emp_id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    employee = conn.execute("SELECT * FROM employees WHERE id = ?", (emp_id,)).fetchone()
    attendance = conn.execute("SELECT * FROM attendance WHERE employee_id = ?", (emp_id,)).fetchall()
    conn.close()
    return render_template('employee_details.html', employee=employee, attendance=attendance)


@employees_bp.route('/employees/cvs')
def employee_cvs():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    employees = conn.execute(
        "SELECT id, name, department, position, document FROM employees WHERE document IS NOT NULL AND document != ''"
    ).fetchall()
    conn.close()
    return render_template('employee_cvs.html', employees=employees)


@employees_bp.route('/employees/cv/<int:employee_id>')
def download_employee_cv(employee_id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    row = conn.execute(
        "SELECT name, document FROM employees WHERE id = ?", (employee_id,)
    ).fetchone()
    conn.close()
    if not row or not row["document"]:
        flash("لا يوجد ملف CV محفوظ لهذا الموظف.", "error")
        return redirect(url_for('employee_cvs'))
    return send_from_directory(UPLOAD_FOLDER, row["document"], as_attachment=True)


@employees_bp.route('/add', methods=['GET', 'POST'])
def add_employee():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            department = request.form.get('department', '').strip()
            position = request.form.get('position', '').strip()
            phone = request.form.get('phone', '').strip()
            email = request.form.get('email', '').strip()
            address = request.form.get('address', '').strip()
            salary = request.form.get('salary', '0')

            # Validate input
            is_valid, msg = validate_employee_data(name, department, position, salary, email, phone)
            if not is_valid:
                flash(msg, "error")
                return redirect(url_for("add_employee"))
            salary = float(salary)
        except Exception as e:
            flash(f"خطأ في البيانات المدخلة: {str(e)}", "error")
            return redirect(url_for("add_employee"))

        portal_username = request.form.get('portal_username', '').strip()
        portal_password = request.form.get('portal_password', '')

        # Handle CV file upload
        cv_file = request.files.get('cv_file') or request.files.get('cv')
        cv_filename = None
        if cv_file and cv_file.filename:
            from app import allowed_cv_file
            is_valid, msg = allowed_cv_file(cv_file)
            if is_valid:
                safe_name = secure_filename(cv_file.filename)
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                cv_filename = f"{timestamp}_{safe_name}"
                cv_file.save(os.path.join(UPLOAD_FOLDER, cv_filename))
            else:
                flash(msg or "صيغة ملف الـ CV غير مدعومة. المسموح: pdf, doc, docx", "error")
                return redirect(url_for("add_employee"))

        username_val = None
        password_hash_val = None
        if portal_username and portal_password:
            username_val = portal_username
            password_hash_val = generate_password_hash(portal_password)

        conn = get_db_connection()
        try:
            conn.execute(
                """INSERT INTO employees (name, department, position, salary, phone, email, document, username, password_hash, address)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (name, department, position, salary, phone, email, cv_filename, username_val, password_hash_val, address)
            )
            conn.commit()
            from app import app
            app.logger.info(f"Employee added: {name}, department: {department}, position: {position}")
        except sqlite3.IntegrityError:
            conn.rollback()
            from app import app
            app.logger.warning(f"Failed to add employee (duplicate): {name}")
            flash("اسم المستخدم للبوابة مُستخدم مسبقاً أو يوجد موظف بنفس الاسم. غيّر الاسم أو اسم المستخدم.", "error")
            return redirect(url_for("add_employee"))
        finally:
            conn.close()

        return redirect(url_for("employees"))

    return render_template("add_employee.html")


@employees_bp.route('/delete/<int:id>')
def delete_employee(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    emp = conn.execute("SELECT name FROM employees WHERE id = ?", (id,)).fetchone()
    emp_name = emp['name'] if emp else f"Unknown id={id}"
    conn.execute("DELETE FROM employees WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    from app import app
    app.logger.warning(f"Employee deleted: {emp_name} (id={id})")
    return redirect('/employees')


@employees_bp.route('/attendance')
def attendance():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    conn = get_db_connection()
    records = conn.execute('''
        SELECT a.id, e.name, a.date, a.check_in, a.check_out
        FROM attendance a
        JOIN employees e ON a.employee_id = e.id
        ORDER BY a.date DESC
    ''').fetchall()
    conn.close()
    return render_template('attendance.html', records=records)


@employees_bp.route('/attendance/checkin/<int:employee_id>', methods=['POST'])
def check_in(employee_id):
    today = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')
    conn = get_db_connection()
    record = conn.execute("SELECT * FROM attendance WHERE employee_id = ? AND date = ?", (employee_id, today)).fetchone()
    if not record:
        conn.execute("INSERT INTO attendance (employee_id, date, check_in) VALUES (?, ?, ?)",
                     (employee_id, today, time_now))
        conn.commit()
    conn.close()
    return redirect('/attendance')


@employees_bp.route('/attendance/checkout/<int:employee_id>', methods=['POST'])
def check_out(employee_id):
    today = datetime.now().strftime('%Y-%m-%d')
    time_now = datetime.now().strftime('%H:%M:%S')
    conn = get_db_connection()
    conn.execute("UPDATE attendance SET check_out = ? WHERE employee_id = ? AND date = ?",
                 (time_now, employee_id, today))
    conn.commit()
    conn.close()
    return redirect('/attendance')


@employees_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_employee(id):
    conn = get_db_connection()

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        position = request.form.get('position', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        salary = request.form.get('salary', '0')
        portal_username = request.form.get('portal_username', '').strip()
        portal_password = request.form.get('portal_password', '')

        # Validate input
        is_valid, msg = validate_employee_data(name, department, position, salary, email, phone)
        if not is_valid:
            flash(msg, "error")
            conn.close()
            return redirect(url_for('edit_employee', id=id))

        try:
            if portal_password:
                password_hash_val = generate_password_hash(portal_password)
                conn.execute(
                    """UPDATE employees SET name=?, department=?, position=?, salary=?, phone=?, email=?, username=?, password_hash=? WHERE id=?""",
                    (name, department, position, float(salary), phone, email, portal_username or None, password_hash_val, id)
                )
            else:
                conn.execute(
                    """UPDATE employees SET name=?, department=?, position=?, salary=?, phone=?, email=?, username=? WHERE id=?""",
                    (name, department, position, float(salary), phone, email, portal_username or None, id)
                )
            conn.commit()
            from app import app
            app.logger.info(f"Employee updated: {name} (id={id})")
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("اسم المستخدم للبوابة مُستخدم من موظف آخر. اختر اسماً مختلفاً.", "error")
            conn.close()
            return redirect(url_for('edit_employee', id=id))
        finally:
            conn.close()
        return redirect(url_for('employees'))

    employee = conn.execute("SELECT * FROM employees WHERE id=?", (id,)).fetchone()
    conn.close()
    return render_template('edit_employee.html', employee=employee)


@employees_bp.route('/download_report')
def download_report():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    conn = get_db_connection()
    data = conn.execute('SELECT * FROM employees ORDER BY name').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = ['ID', 'الاسم', 'الوظيفة', 'القسم', 'الراتب', 'العنوان']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="0b1c8c", end_color="0b1c8c", fill_type="solid")

    for row in data:
        ws.append([row["id"], row["name"], row["department"], row["position"], row["salary"], row["address"]])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_report.xlsx'
    )


@employees_bp.route('/payroll')
def payroll():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    keyword = request.args.get('keyword', '').strip()
    try:
        per_page = int(request.args.get('per_page', 20))
        if per_page not in (10, 20, 50):
            per_page = 20
    except ValueError:
        per_page = 20

    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1

    offset = (page - 1) * per_page

    conn = get_db_connection()

    if keyword:
        like = '%' + keyword + '%'
        payrolls = conn.execute('''
            SELECT payrolls.id, employees.name as employee_name, payrolls.salary,
                   payrolls.bonus, payrolls.deductions, payrolls.net_salary, payrolls.payment_date
            FROM payrolls
            JOIN employees ON payrolls.employee_id = employees.id
            WHERE employees.name LIKE ? OR CAST(payrolls.id AS TEXT) LIKE ?
            ORDER BY payrolls.payment_date DESC
            LIMIT ? OFFSET ?
        ''', (like, like, per_page, offset)).fetchall()

        total = conn.execute('''
            SELECT COUNT(*) as c
            FROM payrolls
            JOIN employees ON payrolls.employee_id = employees.id
            WHERE employees.name LIKE ? OR CAST(payrolls.id AS TEXT) LIKE ?
        ''', (like, like)).fetchone()['c']
    else:
        payrolls = conn.execute('''
            SELECT payrolls.id, employees.name as employee_name, payrolls.salary,
                   payrolls.bonus, payrolls.deductions, payrolls.net_salary, payrolls.payment_date
            FROM payrolls
            JOIN employees ON payrolls.employee_id = employees.id
            ORDER BY payrolls.payment_date DESC
            LIMIT ? OFFSET ?
        ''', (per_page, offset)).fetchall()

        total = conn.execute("SELECT COUNT(*) as c FROM payrolls").fetchone()['c']

    conn.close()

    pages = (total + per_page - 1) // per_page if total else 1

    return render_template('payroll.html',
                           payrolls=payrolls,
                           keyword=keyword,
                           page=page,
                           pages=pages,
                           total=total,
                           per_page=per_page)


@employees_bp.route('/payroll/export/csv')
def payroll_export_csv():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    conn = get_db_connection()

    query = """
        SELECT p.id, e.name AS employee_name, p.salary, p.bonus,
               p.deductions, p.net_salary, p.payment_date
        FROM payrolls p
        JOIN employees e ON p.employee_id = e.id
        ORDER BY p.payment_date DESC, p.id DESC
    """

    rows = conn.execute(query).fetchall()
    conn.close()

    def generate():
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([
            "Payroll ID", "Employee", "Salary", "Bonus",
            "Deductions", "Net Salary", "Payment Date"
        ])
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        for r in rows:
            writer.writerow([
                r["id"],
                r["employee_name"],
                r["salary"],
                r["bonus"],
                r["deductions"],
                r["net_salary"],
                r["payment_date"]
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=payroll_export.csv"
        }
    )


@employees_bp.route('/api/employees/search')
def api_search_employees():
    keyword = request.args.get('keyword', '').strip()
    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1

    per_page = 20
    offset = (page - 1) * per_page

    conn = get_db_connection()

    if keyword:
        like = '%' + keyword + '%'
        employees = conn.execute(
            """SELECT id, name, email, department
               FROM employees
               WHERE name LIKE ? OR COALESCE(email,'') LIKE ? OR COALESCE(department,'') LIKE ?
               ORDER BY name
               LIMIT ? OFFSET ?""",
            (like, like, like, per_page, offset)
        ).fetchall()

        total = conn.execute(
            "SELECT COUNT(*) as c FROM employees WHERE name LIKE ? OR COALESCE(email,'') LIKE ? OR COALESCE(department,'') LIKE ?",
            (like, like, like)
        ).fetchone()['c']
    else:
        employees = conn.execute(
            "SELECT id, name, email, department FROM employees ORDER BY name LIMIT ? OFFSET ?",
            (per_page, offset)
        ).fetchall()

        total = conn.execute("SELECT COUNT(*) as c FROM employees").fetchone()['c']

    conn.close()

    return jsonify({
        'employees': [dict(emp) for emp in employees],
        'page': page,
        'pages': (total + per_page - 1) // per_page if total else 1,
        'total': total,
        'per_page': per_page
    })


@employees_bp.route('/add_payroll', methods=['GET', 'POST'])
def add_payroll():
    conn = get_db_connection()

    # ======================
    # POST (إضافة راتب)
    # ======================
    if request.method == 'POST':
        try:
            employee_id = request.form.get('employee_id')
            month = request.form.get('month', '')

            salary = float(request.form.get('salary', 0))
            bonus = float(request.form.get('bonus', 0))
            deductions = float(request.form.get('deductions', 0))
            payment_date = request.form.get('payment_date')
            net_salary = salary + bonus - deductions

            conn.execute("""
                INSERT INTO payrolls (employee_id, salary, net_salary, bonus, deductions, payment_date)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (employee_id, salary, net_salary, bonus, deductions,payment_date))

            conn.commit()
            flash("تم إضافة الراتب بنجاح", "success")

            return redirect(url_for('employees.add_payroll'))

        except Exception as e:
            conn.rollback()
            flash(f"خطأ في إضافة الراتب: {str(e)}", "error")

    # ======================
    # GET (عرض البيانات)
    # ======================
    keyword = request.args.get('keyword', '').strip()

    per_page = int(request.args.get('per_page', 20))
    page = int(request.args.get('page', 1))
    offset = (page - 1) * per_page

    if keyword:
        like = f"%{keyword}%"
        employees = conn.execute("""
            SELECT id, name, email FROM employees
            WHERE name LIKE ? OR COALESCE(email,'') LIKE ?
            ORDER BY name
            LIMIT ? OFFSET ?
        """, (like, like, per_page, offset)).fetchall()
    else:
        employees = conn.execute("""
            SELECT id, name, email FROM employees
            ORDER BY name
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()

    conn.close()

    return render_template(
        'add_payroll.html',
        employees=employees,
        keyword=keyword,
        per_page=per_page,
        page=page
    )


@employees_bp.route('/get_salary/<int:employee_id>')
def get_employee_salary(employee_id):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT salary FROM employees WHERE id = ?", (employee_id,))
    result = c.fetchone()
    conn.close()

    if result:
        return jsonify({"salary": result[0]})
    else:
        return jsonify({"salary": None})
